#!/usr/bin/env python3
"""Build an Excel/CSV comparison report from OpenRV64 Linux boot logs."""

from __future__ import annotations

import argparse
import csv
import dataclasses
import pathlib
import subprocess
import sys
from collections import defaultdict

from elf_symbols import (
    ElfImage,
    SymbolTable,
    SymbolizedPc,
    load_symbol_tables,
    parse_elf_image,
    symbolize,
)
from linux_boot_signposts import MARKERS, PROGRESS_RE, Signpost, extract


@dataclasses.dataclass(frozen=True)
class RunLog:
    configuration: str
    path: pathlib.Path


@dataclasses.dataclass
class FunctionObservation:
    configuration: str
    symbol: SymbolizedPc
    cycles: int
    instret: int
    pc: int
    source_log: pathlib.Path
    source_line: int
    samples: int = 1


@dataclasses.dataclass(frozen=True)
class RunOutcome:
    configuration: str
    status: str
    first_cycle: int | None
    last_cycle: int | None
    first_instret: int | None
    last_instret: int | None
    prompt_seen: bool
    panic_seen: bool
    simulation_stop_seen: bool
    no_retire_start: int | None
    no_retire_end: int | None
    no_retire_cycles: int
    last_pc: int | None
    source_logs: str


@dataclasses.dataclass(frozen=True)
class Measurement:
    configuration: str
    signpost: Signpost

    @property
    def cycle_low(self) -> int | None:
        if self.signpost.exact_cycles is not None:
            return self.signpost.exact_cycles
        return (
            self.signpost.lower.cycles
            if self.signpost.lower is not None
            else None
        )

    @property
    def cycle_high(self) -> int | None:
        if self.signpost.exact_cycles is not None:
            return self.signpost.exact_cycles
        return (
            self.signpost.upper.cycles
            if self.signpost.upper is not None
            else None
        )

    @property
    def instret_low(self) -> int | None:
        if self.signpost.exact_instret is not None:
            return self.signpost.exact_instret
        return (
            self.signpost.lower.instret
            if self.signpost.lower is not None
            else None
        )

    @property
    def instret_high(self) -> int | None:
        if self.signpost.exact_instret is not None:
            return self.signpost.exact_instret
        return (
            self.signpost.upper.instret
            if self.signpost.upper is not None
            else None
        )

    @staticmethod
    def estimate(low: int | None, high: int | None) -> float | None:
        if low is None or high is None:
            return None
        return (low + high) / 2

    @property
    def cycle_estimate(self) -> float | None:
        return self.estimate(self.cycle_low, self.cycle_high)

    @property
    def instret_estimate(self) -> float | None:
        return self.estimate(self.instret_low, self.instret_high)

    @property
    def precision(self) -> str:
        if self.signpost.exact_cycles is not None:
            return "exact"
        if self.cycle_low is not None and self.cycle_high is not None:
            return "bracketed"
        if self.cycle_low is not None:
            return "lower bound only"
        if self.cycle_high is not None:
            return "upper bound only"
        return "unknown"

    @property
    def pc_low(self) -> int | None:
        if self.signpost.exact_pc is not None:
            return self.signpost.exact_pc
        return self.signpost.lower.pc if self.signpost.lower is not None else None

    @property
    def pc_high(self) -> int | None:
        if self.signpost.exact_pc is not None:
            return self.signpost.exact_pc
        return self.signpost.upper.pc if self.signpost.upper is not None else None


def parse_run(value: str) -> RunLog:
    try:
        configuration, path_text = value.split("=", 1)
    except ValueError as error:
        raise argparse.ArgumentTypeError(
            "--run must be CONFIGURATION=LOG"
        ) from error
    if not configuration:
        raise argparse.ArgumentTypeError("configuration name must not be empty")
    if not path_text:
        raise argparse.ArgumentTypeError("log path must not be empty")
    return RunLog(configuration, pathlib.Path(path_text))


def collect_function_observations(
    run_logs: list[RunLog],
    symbol_tables: list[SymbolTable],
) -> list[FunctionObservation]:
    observations: dict[
        tuple[str, str, str], FunctionObservation
    ] = {}
    configuration_order = list(
        dict.fromkeys(item.configuration for item in run_logs)
    )

    for run_log in run_logs:
        for line_number, line in enumerate(
            run_log.path.read_text(errors="replace").splitlines(),
            start=1,
        ):
            for match in PROGRESS_RE.finditer(line):
                pc = int(match.group("pc"), 16)
                symbol = symbolize(symbol_tables, pc)
                if symbol is None:
                    continue
                key = (
                    run_log.configuration,
                    symbol.image,
                    symbol.function,
                )
                cycles = int(match.group("cycles"))
                instret = int(match.group("instret"))
                previous = observations.get(key)
                if previous is None:
                    observations[key] = FunctionObservation(
                        configuration=run_log.configuration,
                        symbol=symbol,
                        cycles=cycles,
                        instret=instret,
                        pc=pc,
                        source_log=run_log.path,
                        source_line=line_number,
                    )
                else:
                    previous.samples += 1
                    if cycles < previous.cycles:
                        previous.cycles = cycles
                        previous.instret = instret
                        previous.pc = pc
                        previous.symbol = symbol
                        previous.source_log = run_log.path
                        previous.source_line = line_number

    return sorted(
        observations.values(),
        key=lambda item: (
            configuration_order.index(item.configuration),
            item.cycles,
            item.symbol.image,
            item.symbol.function,
        ),
    )


def collect_run_outcomes(
    run_logs: list[RunLog],
) -> list[RunOutcome]:
    grouped: dict[str, list[RunLog]] = defaultdict(list)
    for run_log in run_logs:
        grouped[run_log.configuration].append(run_log)

    outcomes = []
    for configuration, logs in grouped.items():
        progress_by_cycle: dict[int, tuple[int, int]] = {}
        prompt_seen = False
        panic_seen = False
        simulation_stop_seen = False
        for run_log in logs:
            text = run_log.path.read_text(errors="replace")
            prompt_seen |= "openrv64# " in text
            panic_seen |= "Kernel panic" in text
            simulation_stop_seen |= "SIMULATION STOP cycle=" in text
            for line in text.splitlines():
                for match in PROGRESS_RE.finditer(line):
                    progress_by_cycle[int(match.group("cycles"))] = (
                        int(match.group("instret")),
                        int(match.group("pc"), 16),
                    )

        samples = sorted(progress_by_cycle.items())
        longest_start = None
        longest_end = None
        longest_cycles = 0
        run_start = None
        previous_cycle = None
        previous_instret = None
        for cycle, (instret, _pc) in samples:
            if previous_instret is not None and instret == previous_instret:
                if run_start is None:
                    run_start = previous_cycle
                span = cycle - run_start
                if span > longest_cycles:
                    longest_start = run_start
                    longest_end = cycle
                    longest_cycles = span
            else:
                run_start = None
            previous_cycle = cycle
            previous_instret = instret

        if prompt_seen:
            status = "prompt reached"
        elif panic_seen:
            status = "kernel panic"
        elif longest_cycles >= 1_000_000:
            status = "stalled"
        elif simulation_stop_seen:
            status = "completed without prompt"
        else:
            status = "incomplete"

        outcomes.append(
            RunOutcome(
                configuration=configuration,
                status=status,
                first_cycle=samples[0][0] if samples else None,
                last_cycle=samples[-1][0] if samples else None,
                first_instret=samples[0][1][0] if samples else None,
                last_instret=samples[-1][1][0] if samples else None,
                prompt_seen=prompt_seen,
                panic_seen=panic_seen,
                simulation_stop_seen=simulation_stop_seen,
                no_retire_start=longest_start,
                no_retire_end=longest_end,
                no_retire_cycles=longest_cycles,
                last_pc=samples[-1][1][1] if samples else None,
                source_logs="\n".join(str(item.path) for item in logs),
            )
        )
    return outcomes


def collect(
    run_logs: list[RunLog],
) -> tuple[list[str], list[Measurement]]:
    configurations = list(dict.fromkeys(item.configuration for item in run_logs))
    by_key: dict[tuple[str, str], Measurement] = {}

    for run_log in run_logs:
        for signpost in extract(run_log.path):
            key = (run_log.configuration, signpost.marker.name)
            measurement = Measurement(run_log.configuration, signpost)
            previous = by_key.get(key)
            if previous is None:
                by_key[key] = measurement
                continue

            # Prefer an exact record.  Otherwise retain the earliest occurrence
            # across segmented or accidentally overlapping logs.
            if (
                previous.signpost.exact_cycles is None
                and signpost.exact_cycles is not None
            ):
                by_key[key] = measurement
            elif (
                previous.signpost.exact_cycles is None
                and signpost.exact_cycles is None
                and (measurement.cycle_low or sys.maxsize)
                < (previous.cycle_low or sys.maxsize)
            ):
                by_key[key] = measurement

    order = {marker.name: index for index, marker in enumerate(MARKERS)}
    measurements = sorted(
        by_key.values(),
        key=lambda item: (
            configurations.index(item.configuration),
            order[item.signpost.marker.name],
        ),
    )
    return configurations, measurements


def write_csv(
    path: pathlib.Path,
    measurements: list[Measurement],
    symbol_tables: list[SymbolTable],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as output:
        writer = csv.writer(output)
        writer.writerow(
            (
                "configuration",
                "signpost",
                "description",
                "cycle_low",
                "cycle_high",
                "cycle_estimate",
                "instret_low",
                "instret_high",
                "instret_estimate",
                "precision",
                "pc_low",
                "function_low",
                "pc_high",
                "function_high",
                "source_log",
                "source_line",
            )
        )
        for item in measurements:
            symbol_low = (
                symbolize(symbol_tables, item.pc_low)
                if item.pc_low is not None
                else None
            )
            symbol_high = (
                symbolize(symbol_tables, item.pc_high)
                if item.pc_high is not None
                else None
            )
            writer.writerow(
                (
                    item.configuration,
                    item.signpost.marker.name,
                    item.signpost.marker.description,
                    item.cycle_low,
                    item.cycle_high,
                    item.cycle_estimate,
                    item.instret_low,
                    item.instret_high,
                    item.instret_estimate,
                    item.precision,
                    (
                        f"0x{item.pc_low:016x}"
                        if item.pc_low is not None
                        else None
                    ),
                    symbol_low.text() if symbol_low is not None else None,
                    (
                        f"0x{item.pc_high:016x}"
                        if item.pc_high is not None
                        else None
                    ),
                    symbol_high.text() if symbol_high is not None else None,
                    item.signpost.log,
                    item.signpost.line,
                )
            )


def interval_rows(
    configurations: list[str],
    measurements: list[Measurement],
) -> list[tuple[object, ...]]:
    by_configuration: dict[str, list[Measurement]] = defaultdict(list)
    for item in measurements:
        by_configuration[item.configuration].append(item)

    rows: list[tuple[object, ...]] = []
    for configuration in configurations:
        items = by_configuration[configuration]
        for before, after in zip(items, items[1:]):
            if (
                before.cycle_low is None
                or before.cycle_high is None
                or after.cycle_low is None
                or after.cycle_high is None
                or before.instret_low is None
                or before.instret_high is None
                or after.instret_low is None
                or after.instret_high is None
            ):
                cycle_low = cycle_high = None
                instret_low = instret_high = None
                ipc_low = ipc_high = None
            else:
                cycle_low = max(0, after.cycle_low - before.cycle_high)
                cycle_high = max(0, after.cycle_high - before.cycle_low)
                instret_low = max(0, after.instret_low - before.instret_high)
                instret_high = max(0, after.instret_high - before.instret_low)
                ipc_low = (
                    instret_low / cycle_high if cycle_high > 0 else None
                )
                ipc_high = (
                    instret_high / cycle_low if cycle_low > 0 else None
                )
            rows.append(
                (
                    configuration,
                    before.signpost.marker.name,
                    after.signpost.marker.name,
                    cycle_low,
                    cycle_high,
                    instret_low,
                    instret_high,
                    ipc_low,
                    ipc_high,
                )
            )
    return rows


def write_xlsx(
    path: pathlib.Path,
    baseline: str,
    run_logs: list[RunLog],
    configurations: list[str],
    measurements: list[Measurement],
    symbol_tables: list[SymbolTable],
    function_observations: list[FunctionObservation],
    run_outcomes: list[RunOutcome],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as error:
        raise RuntimeError(
            "openpyxl is required to generate the Excel report"
        ) from error

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    readme = workbook.active
    readme.title = "Read Me"
    readme.append(("OpenRV64 Linux boot configuration comparison",))
    readme.append(())
    readme.append(
        (
            "Cycle and retired-instruction values are exact only when the log "
            "contains a PERF SIGNPOST or PERF MILESTONE record.",
        )
    )
    readme.append(
        (
            "Old UART records are bracketed by adjacent progress samples. "
            "Estimates are interval midpoints; low/high columns retain the "
            "measurement uncertainty.",
        )
    )
    readme.append(
        (
            "Blank cells mean the retained log does not contain enough data. "
            "They must not be interpreted as zero.",
        )
    )
    readme.append(
        (
            "Function Landmarks are periodic progress-PC samples, not a "
            "complete call trace. A function absent from that sheet may still "
            "have executed between samples.",
        )
    )
    readme.append(())
    readme.append(("Baseline", baseline))
    readme.append(())
    readme.append(("Configuration", "Source log"))
    for run_log in run_logs:
        readme.append((run_log.configuration, str(run_log.path)))

    outcomes = workbook.create_sheet("Run Outcomes")
    outcome_headers = (
        "Configuration",
        "Status",
        "First cycle",
        "Last sampled cycle",
        "First instret",
        "Last instret",
        "Prompt seen",
        "Panic seen",
        "Simulation stop seen",
        "Longest no-retire start",
        "Longest no-retire end",
        "Longest no-retire cycles",
        "Last sampled PC",
        "Last sampled function",
        "Source logs",
    )
    outcomes.append(outcome_headers)
    for outcome in run_outcomes:
        last_symbol = (
            symbolize(symbol_tables, outcome.last_pc)
            if outcome.last_pc is not None
            else None
        )
        outcomes.append(
            (
                outcome.configuration,
                outcome.status,
                outcome.first_cycle,
                outcome.last_cycle,
                outcome.first_instret,
                outcome.last_instret,
                outcome.prompt_seen,
                outcome.panic_seen,
                outcome.simulation_stop_seen,
                outcome.no_retire_start,
                outcome.no_retire_end,
                outcome.no_retire_cycles,
                (
                    f"0x{outcome.last_pc:016x}"
                    if outcome.last_pc is not None
                    else None
                ),
                last_symbol.text() if last_symbol is not None else None,
                outcome.source_logs,
            )
        )

    milestones = workbook.create_sheet("Milestones")
    milestone_headers = (
        "Configuration",
        "Signpost",
        "Description",
        "Cycle low",
        "Cycle high",
        "Cycle estimate",
        "Cycle +/-",
        "Instret low",
        "Instret high",
        "Instret estimate",
        "Instret +/-",
        "Cumulative IPC estimate",
        "Precision",
        "PC low",
        "Function at low sample",
        "PC high",
        "Function at high sample",
        "Source log",
        "Source line",
    )
    milestones.append(milestone_headers)
    for item in measurements:
        cycle_error = (
            (item.cycle_high - item.cycle_low) / 2
            if item.cycle_low is not None and item.cycle_high is not None
            else None
        )
        instret_error = (
            (item.instret_high - item.instret_low) / 2
            if item.instret_low is not None and item.instret_high is not None
            else None
        )
        cumulative_ipc = (
            item.instret_estimate / item.cycle_estimate
            if item.instret_estimate is not None
            and item.cycle_estimate is not None
            and item.cycle_estimate > 0
            else None
        )
        symbol_low = (
            symbolize(symbol_tables, item.pc_low)
            if item.pc_low is not None
            else None
        )
        symbol_high = (
            symbolize(symbol_tables, item.pc_high)
            if item.pc_high is not None
            else None
        )
        milestones.append(
            (
                item.configuration,
                item.signpost.marker.name,
                item.signpost.marker.description,
                item.cycle_low,
                item.cycle_high,
                item.cycle_estimate,
                cycle_error,
                item.instret_low,
                item.instret_high,
                item.instret_estimate,
                instret_error,
                cumulative_ipc,
                item.precision,
                (
                    f"0x{item.pc_low:016x}"
                    if item.pc_low is not None
                    else None
                ),
                symbol_low.text() if symbol_low is not None else None,
                (
                    f"0x{item.pc_high:016x}"
                    if item.pc_high is not None
                    else None
                ),
                symbol_high.text() if symbol_high is not None else None,
                str(item.signpost.log),
                item.signpost.line,
            )
        )

    functions = workbook.create_sheet("Function Landmarks")
    function_headers = (
        "Configuration",
        "First sampled cycle",
        "Instret at first sample",
        "Image",
        "Function",
        "Function base",
        "Sampled PC",
        "Offset",
        "Progress samples in function",
        "Evidence",
        "Source log",
        "Source line",
    )
    functions.append(function_headers)
    for observation in function_observations:
        functions.append(
            (
                observation.configuration,
                observation.cycles,
                observation.instret,
                observation.symbol.image,
                observation.symbol.function,
                f"0x{observation.symbol.base:016x}",
                f"0x{observation.pc:016x}",
                observation.symbol.offset,
                observation.samples,
                "periodic PC sample; call entry not proven",
                str(observation.source_log),
                observation.source_line,
            )
        )

    intervals = workbook.create_sheet("Intervals")
    interval_headers = (
        "Configuration",
        "From",
        "To",
        "Cycle delta low",
        "Cycle delta high",
        "Instret delta low",
        "Instret delta high",
        "IPC low",
        "IPC high",
    )
    intervals.append(interval_headers)
    for row in interval_rows(configurations, measurements):
        intervals.append(row)

    by_key = {
        (item.configuration, item.signpost.marker.name): item
        for item in measurements
    }
    comparison = workbook.create_sheet("Comparison")
    comparison.append(
        (
            "Signpost",
            "Description",
            *(
                value
                for configuration in configurations
                for value in (
                    f"{configuration} cycles",
                    f"{configuration} instret",
                    (
                        f"{configuration} cycle improvement vs {baseline}"
                        if configuration != baseline
                        else f"{configuration} baseline"
                    ),
                    (
                        f"{configuration} instret delta vs {baseline}"
                        if configuration != baseline
                        else f"{configuration} instret baseline"
                    ),
                )
            ),
        )
    )
    for marker in MARKERS:
        row: list[object] = [marker.name, marker.description]
        baseline_item = by_key.get((baseline, marker.name))
        baseline_cycles = (
            baseline_item.cycle_estimate if baseline_item is not None else None
        )
        baseline_instret = (
            baseline_item.instret_estimate
            if baseline_item is not None
            else None
        )
        for configuration in configurations:
            item = by_key.get((configuration, marker.name))
            cycles = item.cycle_estimate if item is not None else None
            instret = item.instret_estimate if item is not None else None
            if configuration == baseline:
                improvement = 0.0 if cycles is not None else None
            elif (
                cycles is not None
                and baseline_cycles is not None
                and baseline_cycles > 0
            ):
                improvement = (baseline_cycles - cycles) / baseline_cycles
            else:
                improvement = None
            if configuration == baseline:
                instret_delta = 0.0 if instret is not None else None
            elif (
                instret is not None
                and baseline_instret is not None
                and baseline_instret > 0
            ):
                instret_delta = (instret - baseline_instret) / baseline_instret
            else:
                instret_delta = None
            row.extend((cycles, instret, improvement, instret_delta))
        comparison.append(row)

    dark_fill = PatternFill("solid", fgColor="1F4E78")
    light_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in (outcomes, milestones, intervals, comparison, functions):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = dark_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row_index in range(2, sheet.max_row + 1):
            if row_index % 2 == 0:
                for cell in sheet[row_index]:
                    cell.fill = light_fill

    for column in range(4, 13):
        for cells in milestones.iter_cols(
            min_col=column, max_col=column, min_row=2
        ):
            for entry in cells:
                entry.number_format = (
                    "0.000" if column == 12 else "#,##0.0"
                )
    for row in intervals.iter_rows(min_row=2, min_col=4, max_col=9):
        for cell in row[:4]:
            cell.number_format = "#,##0"
        for cell in row[4:]:
            cell.number_format = "0.000"
    for column in range(3, comparison.max_column + 1):
        is_percent = (column - 3) % 4 in (2, 3)
        for cells in comparison.iter_cols(
            min_col=column, max_col=column, min_row=2
        ):
            for entry in cells:
                entry.number_format = "0.0%" if is_percent else "#,##0"

    if intervals.max_row >= 2:
        intervals.conditional_formatting.add(
            f"H2:I{intervals.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )
    if comparison.max_row >= 2:
        for index in range(len(configurations)):
            improvement_column = 5 + index * 4
            instret_delta_column = 6 + index * 4
            improvement_letter = get_column_letter(improvement_column)
            comparison.conditional_formatting.add(
                f"{improvement_letter}2:"
                f"{improvement_letter}{comparison.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="num",
                    mid_value=0,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )
            instret_letter = get_column_letter(instret_delta_column)
            comparison.conditional_formatting.add(
                f"{instret_letter}2:{instret_letter}{comparison.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="num",
                    mid_value=0,
                    mid_color="63BE7B",
                    end_type="max",
                    end_color="F8696B",
                ),
            )

    widths = {
        "Read Me": (34, 105),
        "Run Outcomes": (
            24, 24, 18, 20, 18, 18, 14, 14, 20, 24, 24, 24,
            20, 44, 85,
        ),
        "Milestones": (
            24, 18, 28, 16, 16, 16, 14, 16, 16, 16, 14, 18, 18,
            20, 42, 20, 42, 70, 12,
        ),
        "Intervals": (24, 16, 16, 18, 18, 18, 18, 12, 12),
        "Function Landmarks": (
            24, 18, 20, 14, 38, 20, 20, 12, 18, 38, 70, 12,
        ),
    }
    for sheet_name, column_widths in widths.items():
        sheet = workbook[sheet_name]
        for index, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
    for column in range(1, comparison.max_column + 1):
        comparison.column_dimensions[get_column_letter(column)].width = (
            28 if column == 2 else 22
        )

    readme["A1"].font = Font(size=16, bold=True, color="1F4E78")
    readme.freeze_panes = "A11"

    if len(configurations) > 0 and comparison.max_row > 1:
        cycle_chart = LineChart()
        cycle_chart.title = "Boot milestone cycles by configuration"
        cycle_chart.y_axis.title = "Cycles"
        cycle_chart.x_axis.title = "Milestone"
        cycle_columns = [3 + index * 4 for index in range(len(configurations))]
        for column in cycle_columns:
            cycle_chart.add_data(
                Reference(
                    comparison,
                    min_col=column,
                    max_col=column,
                    min_row=1,
                    max_row=comparison.max_row,
                ),
                titles_from_data=True,
            )
        cycle_chart.set_categories(
            Reference(
                comparison,
                min_col=1,
                min_row=2,
                max_row=comparison.max_row,
            )
        )
        cycle_chart.height = 9
        cycle_chart.width = 18
        comparison.add_chart(cycle_chart, "A13")

        instret_chart = LineChart()
        instret_chart.title = "Retired instructions at boot milestones"
        instret_chart.y_axis.title = "Instructions retired"
        instret_chart.x_axis.title = "Milestone"
        instret_columns = [
            4 + index * 4 for index in range(len(configurations))
        ]
        for column in instret_columns:
            instret_chart.add_data(
                Reference(
                    comparison,
                    min_col=column,
                    max_col=column,
                    min_row=1,
                    max_row=comparison.max_row,
                ),
                titles_from_data=True,
            )
        instret_chart.set_categories(
            Reference(
                comparison,
                min_col=1,
                min_row=2,
                max_row=comparison.max_row,
            )
        )
        instret_chart.height = 9
        instret_chart.width = 18
        comparison.add_chart(instret_chart, "J13")

    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--run",
        action="append",
        type=parse_run,
        required=True,
        help="configuration and source log as CONFIGURATION=LOG; repeatable",
    )
    parser.add_argument(
        "--baseline",
        required=True,
        help="configuration name used as the cycle-comparison baseline",
    )
    parser.add_argument(
        "--xlsx",
        type=pathlib.Path,
        required=True,
        help="output Excel workbook",
    )
    parser.add_argument(
        "--csv",
        type=pathlib.Path,
        required=True,
        help="output normalized CSV",
    )
    parser.add_argument(
        "--elf",
        action="append",
        type=parse_elf_image,
        default=[],
        help="ELF symbol source as IMAGE=PATH; repeatable",
    )
    parser.add_argument(
        "--nm",
        default="riscv64-linux-gnu-nm",
        help="nm executable used to read ELF symbols",
    )
    args = parser.parse_args()

    missing = [item.path for item in args.run if not item.path.is_file()]
    missing.extend(image.path for image in args.elf if not image.path.is_file())
    if missing:
        for path in missing:
            print(f"error: log does not exist: {path}", file=sys.stderr)
        return 2

    configurations, measurements = collect(args.run)
    if args.baseline not in configurations:
        print(
            f"error: baseline is not a --run configuration: {args.baseline}",
            file=sys.stderr,
        )
        return 2

    try:
        symbol_tables = load_symbol_tables(args.elf, args.nm)
    except (OSError, subprocess.CalledProcessError, ValueError) as error:
        print(f"error: could not load ELF symbols: {error}", file=sys.stderr)
        return 2
    function_observations = collect_function_observations(
        args.run,
        symbol_tables,
    )
    run_outcomes = collect_run_outcomes(args.run)
    write_csv(args.csv, measurements, symbol_tables)
    try:
        write_xlsx(
            args.xlsx,
            args.baseline,
            args.run,
            configurations,
            measurements,
            symbol_tables,
            function_observations,
            run_outcomes,
        )
    except RuntimeError as error:
        print(f"error: {error}", file=sys.stderr)
        return 2
    print(
        f"Wrote {args.xlsx} and {args.csv}: "
        f"{len(configurations)} configurations, "
        f"{len(measurements)} measurements"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
